from io import BytesIO
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1F4E79')
DATA_FONT = Font(name='Calibri', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='BDD7EE'),
    right=Side(style='thin', color='BDD7EE'),
    top=Side(style='thin', color='BDD7EE'),
    bottom=Side(style='thin', color='BDD7EE'),
)


def _apply_header_row(ws, row, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def export_distribution_report(report_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_summary_sheet(wb, report_data)
    _build_detail_sheet(wb, report_data)
    _build_asset_sheet(wb, report_data)

    if report_data.get('budget_comparison'):
        _build_budget_comparison_sheet(wb, report_data)

    if report_data.get('yoy_comparison'):
        _build_yoy_sheet(wb, report_data)

    if report_data.get('retained_earnings'):
        _build_retained_earnings_sheet(wb, report_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_summary_sheet(wb, report_data):
    ws = wb.create_sheet('Summary')
    period = report_data['period']
    summary = report_data['summary']

    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'Distribution Report – Summary by Entity'
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Period info
    period_str = f"Period: {period['type'].capitalize()}  |  Year: {period['year']}"
    if period['quarter']:
        period_str += f"  |  Q{period['quarter']}"
    if period['month']:
        from calendar import month_name
        period_str += f"  |  {month_name[period['month']]}"
    ws.merge_cells('A2:E2')
    ws['A2'] = period_str
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='595959')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Totals box
    ws['A3'] = 'Total Distributions:'
    ws['A3'].font = Font(name='Calibri', bold=True, size=11)
    ws['B3'] = float(summary['total_distributions'])
    ws['B3'].number_format = '"$"#,##0.00'
    ws['B3'].font = Font(name='Calibri', bold=True, size=11, color='1F4E79')

    ws['D3'] = 'Entities:'
    ws['D3'].font = Font(name='Calibri', bold=True, size=11)
    ws['E3'] = summary['entity_count']

    ws['D4'] = 'Assets:'
    ws['D4'].font = Font(name='Calibri', bold=True, size=11)
    ws['E4'] = summary['asset_count']

    # Headers
    headers = ['Entity Name', 'Entity Type', 'Total Received', '# Distributions', 'Top Asset']
    _apply_header_row(ws, 6, headers)
    ws.row_dimensions[6].height = 22

    row = 7
    for i, entity in enumerate(sorted(report_data['by_entity'], key=lambda x: -float(x['total_amount']))):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        top_asset = ''
        if entity['by_asset']:
            top = max(entity['by_asset'], key=lambda x: float(x['total_amount']))
            top_asset = top['asset_name']
        data = [
            entity['entity_name'],
            entity['entity_type'].capitalize(),
            float(entity['total_amount']),
            entity['distribution_count'],
            top_asset,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col == 3:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 4:
                cell.alignment = Alignment(horizontal='center')
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A7'


def _build_detail_sheet(wb, report_data):
    ws = wb.create_sheet('Distribution Detail')

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Distribution Detail'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Date', 'Asset', 'Type', 'Entity', 'Amount', 'Ownership %', 'Notes']
    _apply_header_row(ws, 2, headers)

    row = 3
    for i, item in enumerate(sorted(report_data['detail'], key=lambda x: x['distribution_date'])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        data = [
            item['distribution_date'],
            item['asset_name'],
            item['distribution_type'].replace('_', ' ').title(),
            item['entity_name'],
            float(item['amount']),
            float(item['percentage']),
            '',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col == 5:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 6:
                cell.number_format = '0.0000"%"'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A3'


def _build_asset_sheet(wb, report_data):
    ws = wb.create_sheet('Asset Allocations')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Asset Ownership & Allocation Summary'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Asset', 'Asset Type', 'Total Distributed', '# Distributions', 'Entity', 'Entity Share']
    _apply_header_row(ws, 2, headers)

    # Build asset -> entity breakdown from detail
    asset_entity_map = {}
    for item in report_data['detail']:
        aid = item['asset_id']
        eid = item['entity_id']
        if aid not in asset_entity_map:
            asset_entity_map[aid] = {
                'asset_name': item['asset_name'],
                'entities': {},
            }
        if eid not in asset_entity_map[aid]['entities']:
            asset_entity_map[aid]['entities'][eid] = {
                'entity_name': item['entity_name'],
                'amount': Decimal('0'),
            }
        asset_entity_map[aid]['entities'][eid]['amount'] += Decimal(item['amount'])

    asset_totals = {a['asset_id']: a for a in report_data['by_asset']}

    row = 3
    i = 0
    for asset in report_data['by_asset']:
        aid = asset['asset_id']
        entities = asset_entity_map.get(aid, {}).get('entities', {})
        for eid, edata in entities.items():
            fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
            data = [
                asset['asset_name'],
                asset['asset_type'].capitalize(),
                float(asset['total_amount']),
                asset['distribution_count'],
                edata['entity_name'],
                float(edata['amount']),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = DATA_FONT
                cell.fill = fill
                cell.border = THIN_BORDER
                if col in (3, 6):
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            row += 1
            i += 1

    _auto_width(ws)
    ws.freeze_panes = 'A3'


def _build_budget_comparison_sheet(wb, report_data):
    bc = report_data['budget_comparison']
    ws = wb.create_sheet('Budget vs Actual')

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Budget vs Actual — {bc['budget_name']}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Totals row
    ws['A3'] = 'Total Budgeted:'
    ws['A3'].font = Font(name='Calibri', bold=True, size=11)
    ws['B3'] = float(bc['total_budgeted'])
    ws['B3'].number_format = '"$"#,##0.00'
    ws['C3'] = 'Total Actual:'
    ws['C3'].font = Font(name='Calibri', bold=True, size=11)
    ws['D3'] = float(bc['total_actual'])
    ws['D3'].number_format = '"$"#,##0.00'
    ws['E3'] = 'Variance:'
    ws['E3'].font = Font(name='Calibri', bold=True, size=11)
    ws['F3'] = float(bc['total_variance'])
    ws['F3'].number_format = '"$"#,##0.00'
    ws['F3'].font = Font(name='Calibri', bold=True, size=11,
                         color='ED6C02' if float(bc['total_variance']) >= 0 else '2E7D32')

    # By Entity section
    ws['A5'] = 'By Entity'
    ws['A5'].font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
    headers = ['Entity', 'Budgeted', 'Actual', 'Variance', 'Variance %', 'Status']
    _apply_header_row(ws, 6, headers)

    row = 7
    for i, item in enumerate(bc.get('by_entity', [])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        variance = float(item['variance'])
        data = [
            item['entity_name'],
            float(item['budgeted']),
            float(item['actual']),
            variance,
            f"{item['variance_pct']}%" if item['variance_pct'] else 'N/A',
            'Over' if variance >= 0 else 'Under',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 3, 4):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # By Asset section
    row += 1
    ws.cell(row=row, column=1, value='By Asset').font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
    row += 1
    headers = ['Asset', 'Budgeted', 'Actual', 'Variance', 'Variance %', 'Status']
    _apply_header_row(ws, row, headers)
    row += 1

    for i, item in enumerate(bc.get('by_asset', [])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        variance = float(item['variance'])
        data = [
            item['asset_name'],
            float(item['budgeted']),
            float(item['actual']),
            variance,
            f"{item['variance_pct']}%" if item['variance_pct'] else 'N/A',
            'Over' if variance >= 0 else 'Under',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 3, 4):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A7'


def _build_yoy_sheet(wb, report_data):
    yoy = report_data['yoy_comparison']
    ws = wb.create_sheet('Year over Year')

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Year-over-Year Comparison — {yoy['prior_year']} vs {yoy['current_year']}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Totals
    ws['A3'] = f"{yoy['prior_year']} Total:"
    ws['A3'].font = Font(name='Calibri', bold=True, size=11)
    ws['B3'] = float(yoy['total_prior'])
    ws['B3'].number_format = '"$"#,##0.00'
    ws['C3'] = f"{yoy['current_year']} Total:"
    ws['C3'].font = Font(name='Calibri', bold=True, size=11)
    ws['D3'] = float(yoy['total_current'])
    ws['D3'].number_format = '"$"#,##0.00'
    ws['E3'] = 'Change:'
    ws['E3'].font = Font(name='Calibri', bold=True, size=11)
    ws['F3'] = float(yoy['total_change'])
    ws['F3'].number_format = '"$"#,##0.00'

    # By Entity
    ws['A5'] = 'By Entity'
    ws['A5'].font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
    headers = ['Entity', f'{yoy["prior_year"]}', f'{yoy["current_year"]}', 'Change ($)', 'Change (%)', 'Trend']
    _apply_header_row(ws, 6, headers)

    row = 7
    for i, item in enumerate(yoy.get('by_entity', [])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        change = float(item['change'])
        data = [
            item['entity_name'],
            float(item['prior_amount']),
            float(item['current_amount']),
            change,
            f"{item['change_pct']}%" if item['change_pct'] else 'N/A',
            '↑' if change > 0 else ('↓' if change < 0 else '→'),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 3, 4):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # By Asset
    row += 1
    ws.cell(row=row, column=1, value='By Asset').font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
    row += 1
    headers = ['Asset', f'{yoy["prior_year"]}', f'{yoy["current_year"]}', 'Change ($)', 'Change (%)', 'Trend']
    _apply_header_row(ws, row, headers)
    row += 1

    for i, item in enumerate(yoy.get('by_asset', [])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        change = float(item['change'])
        data = [
            item['asset_name'],
            float(item['prior_amount']),
            float(item['current_amount']),
            change,
            f"{item['change_pct']}%" if item['change_pct'] else 'N/A',
            '↑' if change > 0 else ('↓' if change < 0 else '→'),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 3, 4):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A7'


def _build_retained_earnings_sheet(wb, report_data):
    re_data = report_data['retained_earnings']
    ws = wb.create_sheet('Retained Earnings')

    ws.merge_cells('A1:E1')
    ws['A1'] = f"Retained Earnings Rollforward — {re_data['year']}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Summary totals
    ws['A3'] = 'Beginning Balance:'
    ws['A3'].font = Font(name='Calibri', bold=True, size=11)
    ws['B3'] = float(re_data['total_beginning_balance'])
    ws['B3'].number_format = '"$"#,##0.00'
    ws['C3'] = f'{re_data["year"]} Distributions:'
    ws['C3'].font = Font(name='Calibri', bold=True, size=11)
    ws['D3'] = float(re_data['total_current_year'])
    ws['D3'].number_format = '"$"#,##0.00'

    ws['A4'] = 'Ending Balance:'
    ws['A4'].font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    ws['B4'] = float(re_data['total_ending_balance'])
    ws['B4'].number_format = '"$"#,##0.00'
    ws['B4'].font = Font(name='Calibri', bold=True, size=11, color='1F4E79')

    headers = ['Entity', 'Beginning Balance', f'{re_data["year"]} Distributions', 'Ending Balance']
    _apply_header_row(ws, 6, headers)

    row = 7
    for i, item in enumerate(re_data.get('by_entity', [])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        data = [
            item['entity_name'],
            float(item['beginning_balance']),
            float(item['current_year_distributions']),
            float(item['ending_balance']),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 3, 4):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Totals row
    fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    bold = Font(name='Calibri', bold=True, size=10)
    totals = [
        'Total',
        float(re_data['total_beginning_balance']),
        float(re_data['total_current_year']),
        float(re_data['total_ending_balance']),
    ]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = bold
        cell.fill = fill
        cell.border = THIN_BORDER
        if col in (2, 3, 4):
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal='right')

    _auto_width(ws)
    ws.freeze_panes = 'A7'


# ═══════════════════════════════════════════════════════════════════════
# FMV Report Export
# ═══════════════════════════════════════════════════════════════════════

def export_fmv_report(report_data):
    """Export FMV report to Excel with Summary and Line Items sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_fmv_summary_sheet(wb, report_data)
    _build_fmv_line_items_sheet(wb, report_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_fmv_summary_sheet(wb, report_data):
    """Build the FMV Summary sheet."""
    ws = wb.create_sheet('Summary')

    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'FMV Report'
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Date
    from datetime import date
    ws.merge_cells('A2:E2')
    ws['A2'].value = f'Generated: {date.today().isoformat()}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Total FMV
    ws['A4'].value = 'Total FMV'
    ws['A4'].font = Font(name='Calibri', bold=True, size=12)
    ws['B4'].value = float(report_data['total_fmv'])
    ws['B4'].font = Font(name='Calibri', bold=True, size=12, color='2E7D32')
    ws['B4'].number_format = '"$"#,##0.00'

    # Applied filters
    filters = report_data.get('filters', {})
    type_filters = filters.get('type_filters', [])
    entity_ids = filters.get('entity_ids', [])
    ws['A5'].value = 'Filters'
    ws['A5'].font = Font(name='Calibri', bold=True, size=10)
    filter_text = 'None'
    filter_parts = []
    if type_filters:
        filter_parts.append(f"Types: {', '.join(type_filters)}")
    if entity_ids:
        filter_parts.append(f"Entities: {', '.join(str(e) for e in entity_ids)}")
    if filter_parts:
        filter_text = '; '.join(filter_parts)
    ws['B5'].value = filter_text
    ws['B5'].font = DATA_FONT

    # Type breakdown table
    row = 7
    headers = ['Asset Type', 'Total Value', 'Items', 'Allocation %']
    _apply_header_row(ws, row, headers)
    row += 1

    for i, type_data in enumerate(report_data.get('by_type', [])):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row=row, column=1, value=type_data['label']).font = DATA_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = THIN_BORDER

        val_cell = ws.cell(row=row, column=2, value=float(type_data['total_value']))
        val_cell.font = DATA_FONT
        val_cell.number_format = '"$"#,##0.00'
        val_cell.alignment = Alignment(horizontal='right')
        val_cell.fill = fill
        val_cell.border = THIN_BORDER

        count_cell = ws.cell(row=row, column=3, value=type_data['count'])
        count_cell.font = DATA_FONT
        count_cell.alignment = Alignment(horizontal='center')
        count_cell.fill = fill
        count_cell.border = THIN_BORDER

        pct_cell = ws.cell(row=row, column=4, value=float(type_data['percentage']))
        pct_cell.font = DATA_FONT
        pct_cell.number_format = '0.00"%"'
        pct_cell.alignment = Alignment(horizontal='right')
        pct_cell.fill = fill
        pct_cell.border = THIN_BORDER

        row += 1

    _auto_width(ws)


def _build_fmv_line_items_sheet(wb, report_data):
    """Build the FMV Line Items sheet."""
    ws = wb.create_sheet('Line Items')

    headers = ['Name', 'Value', 'Source', 'Asset Type', 'Institution', 'Subtype', 'Snapshot Date']
    _apply_header_row(ws, 1, headers)

    for i, item in enumerate(report_data.get('items', [])):
        row = i + 2
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL

        ws.cell(row=row, column=1, value=item['name']).font = DATA_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = THIN_BORDER

        val_cell = ws.cell(row=row, column=2, value=float(item['value']))
        val_cell.font = DATA_FONT
        val_cell.number_format = '"$"#,##0.00'
        val_cell.alignment = Alignment(horizontal='right')
        val_cell.fill = fill
        val_cell.border = THIN_BORDER

        source_cell = ws.cell(row=row, column=3, value=item['source'].capitalize())
        source_cell.font = DATA_FONT
        source_cell.fill = fill
        source_cell.border = THIN_BORDER

        type_cell = ws.cell(row=row, column=4, value=item['label'])
        type_cell.font = DATA_FONT
        type_cell.fill = fill
        type_cell.border = THIN_BORDER

        inst_cell = ws.cell(row=row, column=5, value=item.get('institution') or '')
        inst_cell.font = DATA_FONT
        inst_cell.fill = fill
        inst_cell.border = THIN_BORDER

        sub_cell = ws.cell(row=row, column=6, value=item.get('subtype') or '')
        sub_cell.font = DATA_FONT
        sub_cell.fill = fill
        sub_cell.border = THIN_BORDER

        date_cell = ws.cell(row=row, column=7, value=item.get('snapshot_date') or '')
        date_cell.font = DATA_FONT
        date_cell.fill = fill
        date_cell.border = THIN_BORDER

    _auto_width(ws)
    ws.freeze_panes = 'A2'


# ---------------------------------------------------------------------------
# Portfolio Tracker Excel Exports
# ---------------------------------------------------------------------------

def _fmt_money(val):
    """Format a string-encoded Decimal as a float for Excel, or 0."""
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def _fmt_ratio(val):
    """Format ratio field — None → 'N/A', else float."""
    if val is None:
        return 'N/A'
    try:
        return float(val)
    except (ValueError, TypeError):
        return 'N/A'


def _fmt_pct(val):
    """Format percentage field — None → '—', else float string with %."""
    if val is None:
        return '—'
    try:
        return float(val)
    except (ValueError, TypeError):
        return '—'


MONEY_FMT = '#,##0.00'
PCT_FMT = '0.00"%"'
RATIO_FMT = '0.00'


def export_portfolio_summary(report_data):
    """Export Portfolio Summary (entity rollups) to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portfolio Summary'

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Portfolio Summary — Entity Rollups (as of {report_data.get('as_of_date', '')})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        'Entity', 'Original Commitment', '% Called', 'Unfunded',
        'Paid-In', 'Distributions', 'Residual Value',
        'DPI', 'RVPI', 'TVPI', 'IRR (%)',
    ]
    _apply_header_row(ws, 2, headers)

    # Entity rows
    row = 3
    for ent in report_data.get('entities', []):
        fill = ALT_ROW_FILL if (row % 2 == 1) else WHITE_FILL

        cells_data = [
            (ent.get('entity_name', ''), None),
            (_fmt_money(ent.get('original_commitment')), MONEY_FMT),
            (_fmt_pct(ent.get('pct_called')), None),
            (_fmt_money(ent.get('unfunded_commitment')), MONEY_FMT),
            (_fmt_money(ent.get('paid_in')), MONEY_FMT),
            (_fmt_money(ent.get('distributions')), MONEY_FMT),
            (_fmt_money(ent.get('residual')), MONEY_FMT),
            (_fmt_ratio(ent.get('dpi')), None),
            (_fmt_ratio(ent.get('rvpi')), None),
            (_fmt_ratio(ent.get('tvpi')), None),
            (_fmt_ratio(ent.get('irr')), None),
        ]
        for col, (val, num_fmt) in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if num_fmt:
                cell.number_format = num_fmt
            if col > 1:
                cell.alignment = Alignment(horizontal='right')

        row += 1

    # All Entities total row
    totals = report_data.get('all_entities')
    if totals:
        total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_font = Font(name='Calibri', bold=True, size=10)
        total_data = [
            ('All Entities', None),
            (_fmt_money(totals.get('original_commitment')), MONEY_FMT),
            (_fmt_pct(totals.get('pct_called')), None),
            (_fmt_money(totals.get('unfunded_commitment')), MONEY_FMT),
            (_fmt_money(totals.get('paid_in')), MONEY_FMT),
            (_fmt_money(totals.get('distributions')), MONEY_FMT),
            (_fmt_money(totals.get('residual')), MONEY_FMT),
            (_fmt_ratio(totals.get('dpi')), None),
            (_fmt_ratio(totals.get('rvpi')), None),
            (_fmt_ratio(totals.get('tvpi')), None),
            (_fmt_ratio(totals.get('irr')), None),
        ]
        for col, (val, num_fmt) in enumerate(total_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = THIN_BORDER
            if num_fmt:
                cell.number_format = num_fmt
            if col > 1:
                cell.alignment = Alignment(horizontal='right')

    _auto_width(ws)
    ws.freeze_panes = 'A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_asset_class_summary(report_data):
    """Export Asset Class Summary to an Excel workbook matching the spreadsheet format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asset Class Summary'

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Asset Class Summary — Rollups by Asset Class (as of {report_data.get('as_of_date', '')})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers matching spreadsheet
    headers = [
        'Asset Class', 'Original Commitment', '% Called', 'Unfunded Commitment',
        'Paid-In (ABS)', 'Distributions', 'Residual Used',
        'DPI', 'RVPI', 'TVPI', 'IRR (XIRR)',
    ]
    _apply_header_row(ws, 3, headers)

    def _dash_or_money(val):
        """Return formatted money or ' - ' for zero/None."""
        v = _fmt_money(val)
        return ' - ' if v == 0 or v is None else v

    def _dash_or_ratio(val):
        if val is None:
            return ' - '
        v = float(val)
        return ' - ' if v == 0 else v

    def _dash_or_pct(val):
        if val is None:
            return ''
        v = float(val)
        return f'{v:.0f}%' if v > 0 else ''

    def _dash_or_irr(val):
        if val is None:
            return ''
        return float(val) / 100  # stored as percentage number

    row = 4
    for cls in report_data.get('rows', []):
        fill = ALT_ROW_FILL if (row % 2 == 0) else WHITE_FILL

        values = [
            cls.get('asset_class', ''),
            _dash_or_money(cls.get('original_commitment')),
            _dash_or_pct(cls.get('pct_called')),
            _dash_or_money(cls.get('unfunded_commitment')),
            _dash_or_money(cls.get('paid_in')),
            _dash_or_money(cls.get('distributions')),
            _dash_or_money(cls.get('residual')),
            _dash_or_ratio(cls.get('dpi')),
            _dash_or_ratio(cls.get('rvpi')),
            _dash_or_ratio(cls.get('tvpi')),
            _dash_or_irr(cls.get('irr')),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in (2, 4, 5, 6, 7):  # money columns
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal='right')
            elif col in (8, 9, 10):  # ratio columns
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col == 11:  # IRR
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='right')
            elif col == 3:  # pct called
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Blank row then All Asset Classes total
    row += 1
    ac = report_data.get('all_classes', {})
    total_values = [
        'All Asset Classes',
        _dash_or_money(ac.get('original_commitment')),
        _dash_or_pct(ac.get('pct_called')),
        _dash_or_money(ac.get('unfunded_commitment')),
        _dash_or_money(ac.get('paid_in')),
        _dash_or_money(ac.get('distributions')),
        _dash_or_money(ac.get('residual')),
        _dash_or_ratio(ac.get('dpi')),
        _dash_or_ratio(ac.get('rvpi')),
        _dash_or_ratio(ac.get('tvpi')),
        _dash_or_irr(ac.get('irr')),
    ]
    for col, val in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        cell.border = THIN_BORDER
        if col in (2, 4, 5, 6, 7):
            cell.number_format = MONEY_FMT
            cell.alignment = Alignment(horizontal='right')
        elif col in (8, 9, 10):
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='right')
        elif col == 11:
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='right')

    _auto_width(ws)
    ws.freeze_panes = 'A4'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_investment_performance(report_data):
    """Export Investment Performance to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Investment Performance'

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Investment Performance (as of {report_data.get('as_of_date', '')})"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Per-investment headers
    headers = [
        'Investment', 'Entity', 'Type', 'Commitment',
        'Paid-In', 'Distributions', 'Residual',
        'DPI', 'RVPI', 'TVPI', 'IRR (%)',
    ]
    _apply_header_row(ws, 2, headers)

    row = 3
    for inv in report_data.get('investments', []):
        fill = ALT_ROW_FILL if (row % 2 == 1) else WHITE_FILL
        cells_data = [
            (inv.get('asset_name', ''), None),
            (inv.get('entity_name', ''), None),
            (inv.get('asset_type', ''), None),
            (_fmt_money(inv.get('original_commitment')), MONEY_FMT),
            (_fmt_money(inv.get('paid_in')), MONEY_FMT),
            (_fmt_money(inv.get('distributions')), MONEY_FMT),
            (_fmt_money(inv.get('residual')), MONEY_FMT),
            (_fmt_ratio(inv.get('dpi')), None),
            (_fmt_ratio(inv.get('rvpi')), None),
            (_fmt_ratio(inv.get('tvpi')), None),
            (_fmt_ratio(inv.get('irr')), None),
        ]
        for col, (val, num_fmt) in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if num_fmt:
                cell.number_format = num_fmt
            if col > 3:
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Entity totals section
    entity_totals = report_data.get('entity_totals', [])
    if entity_totals:
        row += 1  # blank separator
        total_headers = [
            'Entity', 'Paid-In', 'Distributions', 'Residual',
            'DPI', 'RVPI', 'TVPI', 'IRR (%)',
        ]
        _apply_header_row(ws, row, total_headers)
        row += 1

        total_font = Font(name='Calibri', bold=True, size=10)
        for et in entity_totals:
            fill = ALT_ROW_FILL if (row % 2 == 1) else WHITE_FILL
            et_data = [
                (et.get('entity_name', ''), None),
                (_fmt_money(et.get('paid_in')), MONEY_FMT),
                (_fmt_money(et.get('distributions')), MONEY_FMT),
                (_fmt_money(et.get('residual')), MONEY_FMT),
                (_fmt_ratio(et.get('dpi')), None),
                (_fmt_ratio(et.get('rvpi')), None),
                (_fmt_ratio(et.get('tvpi')), None),
                (_fmt_ratio(et.get('irr')), None),
            ]
            for col, (val, num_fmt) in enumerate(et_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = total_font
                cell.fill = fill
                cell.border = THIN_BORDER
                if num_fmt:
                    cell.number_format = num_fmt
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
            row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# K-1 Document Export
# ---------------------------------------------------------------------------

def export_k1_summary(k1_documents):
    """Export a list of K-1 documents to an Excel workbook.

    Args:
        k1_documents: QuerySet of K1Document instances with related objects.

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'K-1 Summary'

    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'K-1 Document Summary'
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        'Tax Year', 'Partnership', 'EIN', 'Status',
        'Entity', 'Net Income', 'Distributions', 'Ending Capital',
    ]
    _apply_header_row(ws, 2, headers)

    row = 3
    for doc in k1_documents:
        pi = getattr(doc, 'partnership_info', None)
        cap = getattr(doc, 'capital_account', None)
        dist_items = doc.income_items.filter(line_number='19')
        total_dist = sum(i.amount or 0 for i in dist_items)

        values = [
            doc.tax_year,
            pi.name if pi else doc.original_filename,
            pi.ein if pi else '',
            doc.status.capitalize(),
            doc.entity.name if doc.entity else '',
            cap.net_income if cap else None,
            total_dist if total_dist else None,
            cap.ending_balance if cap else None,
        ]

        fill = ALT_ROW_FILL if row % 2 == 1 else WHITE_FILL
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col >= 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════
# Activity Report Export
# ═══════════════════════════════════════════════════════════════════════

def export_activity_report(activities):
    """Export a list of Activity model instances to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Activity'

    # Title
    ws.merge_cells('A1:V1')
    title_cell = ws['A1']
    title_cell.value = 'Activity — All Entities & Partnerships'
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        'Year', 'Entity', 'Partnership',
        'Beg Basis', 'Contributions',
        'Interest (5)', 'Dividends (6)', 'Cap Gains (8/9/10)',
        'Remaining K-1 Income/Ded.', 'Total Income',
        'Distributions', 'Other Adj (18-c)',
        'Ending Tax Basis', 'Ending GL Balance Per Books',
        'Book-to-Tax Adj', 'Ending K-1 Capital Account',
        'K-1 Capital vs Tax Basis Diff', 'Excess Distribution',
        'Negative Basis?', 'Δ Ending Basis vs Prior Year',
        'Notes', 'Source',
    ]
    _apply_header_row(ws, 2, headers)

    # Data
    row = 3
    money_cols = set(range(4, 19))  # columns D through R (1-indexed)
    money_cols.add(20)  # T = Δ Ending Basis

    for act in activities:
        vals = [
            act.year,
            act.entity.name if act.entity else '',
            act.asset.name if act.asset else '',
            float(act.beginning_basis),
            float(act.contributions),
            float(act.interest),
            float(act.dividends),
            float(act.capital_gains),
            float(act.remaining_k1_income),
            float(act.total_income),
            float(act.distributions),
            float(act.other_adjustments),
            float(act.ending_tax_basis),
            float(act.ending_gl_balance),
            float(act.book_to_tax_adj),
            float(act.ending_k1_capital),
            float(act.k1_capital_vs_tax_diff),
            float(act.excess_distribution),
            'YES' if act.negative_basis else '',
            float(act.basis_change),
            act.notes or '',
            f"K-1 {act.source_k1_document.tax_year}" if act.source_k1_document else 'Manual',
        ]
        fill = ALT_ROW_FILL if (row - 3) % 2 == 1 else WHITE_FILL
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col in money_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'D3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf